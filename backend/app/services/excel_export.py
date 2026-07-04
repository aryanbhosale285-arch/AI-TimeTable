"""Build an .xlsx workbook from a timetable — one sheet per class.

Each sheet is a day × period grid with the subject + teacher in each cell,
colour-coded by subject. Kept dependency-light (openpyxl only).
"""
from io import BytesIO
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill(start_color="F1F5F9", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _argb(hex_color: str | None) -> str | None:
    if not hex_color:
        return None
    h = hex_color.lstrip("#")
    if len(h) == 6:
        return "FF" + h.upper()
    return None


def _render_sheet(ws, days: list, periods: int, period_labels: dict, grid: dict) -> None:
    """Draw one day × period grid. Cell info: {"subject","teacher","color"}."""
    ws.column_dimensions["A"].width = 14

    hc = ws.cell(1, 1, "Day / Period")
    hc.font = Font(bold=True)
    hc.fill = _HEADER_FILL
    hc.border = _BORDER
    hc.alignment = _CENTER
    for p in range(periods):
        label = f"P{p + 1}"
        if period_labels.get(p):
            label += f"\n{period_labels[p]}"
        c = ws.cell(1, p + 2, label)
        c.font = Font(bold=True)
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _CENTER
        ws.column_dimensions[c.column_letter].width = 16

    for di, day_name in enumerate(days):
        dc = ws.cell(di + 2, 1, day_name)
        dc.font = Font(bold=True)
        dc.fill = _HEADER_FILL
        dc.border = _BORDER
        dc.alignment = _CENTER
        for p in range(periods):
            cell = ws.cell(di + 2, p + 2)
            cell.border = _BORDER
            cell.alignment = _CENTER
            info = grid.get((di, p))
            if info and info.get("subject"):
                text = info["subject"]
                if info.get("teacher"):
                    text += f"\n{info['teacher']}"
                cell.value = text
                argb = _argb(info.get("color"))
                if argb:
                    cell.fill = PatternFill(start_color=argb, fill_type="solid")
        ws.row_dimensions[di + 2].height = 34


def _sheet_name(label: str, used: set) -> str:
    """Excel-safe unique sheet name (31 chars, no []:*?/\\)."""
    clean = "".join(ch for ch in label if ch not in "[]:*?/\\")[:31] or "Sheet"
    name, n = clean, 2
    while name in used:
        suffix = f" ({n})"
        name = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def build_timetable_xlsx(
    *,
    days: list,            # ordered day names
    periods: int,
    sections: list,        # [{"id": int, "label": str}]
    period_labels: dict,   # period_index -> "08:00–08:45" or None
    slots_by_section: dict,  # section_id -> {(day_idx, period_idx): {"subject","teacher","color"}}
    title: str,
    teachers: list | None = None,     # [{"id": int, "label": str}] for the full-school bundle
    slots_by_teacher: dict | None = None,  # teacher_id -> same grid shape; "teacher" holds the class label
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()

    for sec in sections:
        ws = wb.create_sheet(title=_sheet_name(sec["label"] or f"Class {sec['id']}", used_names))
        _render_sheet(ws, days, periods, period_labels, slots_by_section.get(sec["id"], {}))

    # Full-school bundle: one sheet per teacher after the class sheets.
    for t in teachers or []:
        ws = wb.create_sheet(title=_sheet_name(f"T - {t['label']}", used_names))
        _render_sheet(ws, days, periods, period_labels, (slots_by_teacher or {}).get(t["id"], {}))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
