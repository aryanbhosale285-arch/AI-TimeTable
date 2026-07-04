"""End-to-end API smoke test against a throwaway SQLite database.

Covers the production-readiness surface added in v2.1:
auth (register/login/401s), school ownership, CSV import, background
generation jobs, share links (create/resolve/revoke), and the Excel bundle.

Run:  python test_api_smoke.py
"""
import os
import sys
import time

TEST_DB = "test_smoke.db"
if os.path.exists(TEST_DB):
    os.remove(TEST_DB)
os.environ["DATABASE_URL"] = f"sqlite:///./{TEST_DB}"

from fastapi.testclient import TestClient  # noqa: E402
from app.main import app  # noqa: E402

client = TestClient(app)
FAILURES = []


def check(name: str, cond: bool, detail: str = ""):
    status = "PASS" if cond else "FAIL"
    print(f"[{status}] {name}" + (f" — {detail}" if detail and not cond else ""))
    if not cond:
        FAILURES.append(name)


# ---- auth ----
r = client.get("/api/schools")
check("unauthenticated /schools is rejected", r.status_code == 401, str(r.status_code))

r = client.post("/api/auth/register", json={"email": "admin@test.io", "name": "Admin", "password": "supersecret1"})
check("register", r.status_code == 201, r.text)
token = r.json()["access_token"]
HDRS = {"Authorization": f"Bearer {token}"}

r = client.post("/api/auth/register", json={"email": "admin@test.io", "name": "Dup", "password": "supersecret1"})
check("duplicate email rejected", r.status_code == 409, r.text)

r = client.post("/api/auth/login", json={"email": "admin@test.io", "password": "wrongpass99"})
check("wrong password rejected", r.status_code == 401, r.text)

r = client.post("/api/auth/login", json={"email": "admin@test.io", "password": "supersecret1"})
check("login", r.status_code == 200, r.text)

r = client.get("/api/auth/me", headers=HDRS)
check("auth/me", r.status_code == 200 and r.json()["email"] == "admin@test.io", r.text)

# ---- school setup ----
school_body = {
    "name": "Smoke Test School",
    "board": "CBSE",
    "periods_per_day": 8,
    "half_day_periods": None,
    "academic_year": "2026-27",
    "working_days": [
        {"day_name": d, "is_half_day": False, "day_order": i}
        for i, d in enumerate(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"])
    ],
    "periods": [
        {"period_number": i + 1, "start_time": f"{8 + i:02d}:00", "end_time": f"{8 + i:02d}:45"}
        for i in range(8)
    ],
    "breaks": [{"name": "Lunch", "after_period": 4, "duration_minutes": 30}],
}
r = client.post("/api/schools", json=school_body, headers=HDRS)
check("create school", r.status_code == 200, r.text)
sid = r.json()["id"]

r = client.get("/api/schools", headers=HDRS)
check("list schools shows owned school", any(s["id"] == sid for s in r.json()), r.text)

# second user cannot touch the first user's school
r2 = client.post("/api/auth/register", json={"email": "other@test.io", "name": "Other", "password": "supersecret2"})
other = {"Authorization": f"Bearer {r2.json()['access_token']}"}
r = client.get(f"/api/schools/{sid}", headers=other)
check("other user forbidden from owned school", r.status_code == 403, str(r.status_code))
r = client.get(f"/api/schools/{sid}/teachers", headers=other)
check("school-scoped routes also forbidden", r.status_code == 403, str(r.status_code))

# ---- CSV import ----
with open("../teacher_assignments.csv", "rb") as f:
    r = client.post(
        f"/api/schools/{sid}/import",
        files={"file": ("teacher_assignments.csv", f, "text/csv")},
        headers=HDRS,
    )
check("CSV import", r.status_code == 200, r.text)
imported = r.json().get("imported", 0)
check("assignments imported", imported > 0, str(imported))

# ---- background generation job ----
r = client.post(
    f"/api/schools/{sid}/timetables/generate-async",
    json={"name": "Smoke TT", "fixed_slots": []},
    headers=HDRS,
)
check("start async generation", r.status_code == 202, r.text)
job_id = r.json()["id"]

status, stages = "QUEUED", set()
for _ in range(120):
    r = client.get(f"/api/schools/{sid}/timetables/jobs/{job_id}", headers=HDRS)
    job = r.json()
    status = job["status"]
    if job.get("stage"):
        stages.add(job["stage"])
    if status in ("SUCCEEDED", "FAILED"):
        break
    time.sleep(1)
check("job finished", status in ("SUCCEEDED", "FAILED"), status)
check("job succeeded", status == "SUCCEEDED", str(job.get("errors")))
tid = job["timetable_id"]

r = client.get(f"/api/schools/{sid}/timetables/{tid}", headers=HDRS)
check("fetch generated timetable", r.status_code == 200 and len(r.json()["slots"]) > 0, r.text[:200])

# no teacher or class double-booked
slots = r.json()["slots"]
seen_teacher, seen_class, clash = set(), set(), 0
for s in slots:
    if s["teacher_id"]:
        k = (s["teacher_id"], s["day_index"], s["period_index"])
        clash += k in seen_teacher
        seen_teacher.add(k)
    k = (s["section_id"], s["day_index"], s["period_index"])
    clash += k in seen_class
    seen_class.add(k)
check("zero clashes in generated timetable", clash == 0, f"{clash} clashes")

# ---- Excel bundle (class + teacher sheets) ----
r = client.get(f"/api/schools/{sid}/timetables/{tid}/export.xlsx", headers=HDRS)
check("Excel export", r.status_code == 200 and len(r.content) > 1000, str(r.status_code))
from io import BytesIO
from openpyxl import load_workbook
wb = load_workbook(BytesIO(r.content))
teacher_sheets = [n for n in wb.sheetnames if n.startswith("T - ")]
class_sheets = [n for n in wb.sheetnames if not n.startswith("T - ")]
check("bundle has class sheets", len(class_sheets) > 0, str(wb.sheetnames))
check("bundle has teacher sheets", len(teacher_sheets) > 0, str(wb.sheetnames))

# ---- share links ----
r = client.post(f"/api/schools/{sid}/timetables/{tid}/share-links", headers=HDRS)
check("create share link", r.status_code == 201, r.text)
share_token = r.json()["token"]
link_id = r.json()["id"]

r = client.get(f"/api/share/{share_token}")  # no auth
share = r.json()
check("public share resolves without auth", r.status_code == 200, r.text[:200])
check("share hides staff details", all("teacher_name" not in s and "teacher_id" not in s for s in share["slots"]))
check("share includes school + standards", bool(share["school"]["working_days"]) and bool(share["standards"]))

r = client.delete(f"/api/schools/{sid}/timetables/{tid}/share-links/{link_id}", headers=HDRS)
check("revoke share link", r.status_code == 204, str(r.status_code))
r = client.get(f"/api/share/{share_token}")
check("revoked link is dead", r.status_code == 404, str(r.status_code))

# ---- sync generate still works (compat) ----
r = client.post(
    f"/api/schools/{sid}/timetables/generate",
    json={"name": "Sync TT", "fixed_slots": []},
    headers=HDRS,
)
check("sync generate (compat)", r.status_code == 200, r.text[:200])

print()
if FAILURES:
    print(f"[FAIL] {len(FAILURES)} check(s) failed: {FAILURES}")
    sys.exit(1)
print("[PASS] ALL API SMOKE CHECKS PASSED")
